#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
AH Poultry Antibody Titer Dashboard
Lab Report - Data Extraction Utility

A user-friendly dashboard for extracting structured data from veterinary lab reports
using LLM-powered OCR and providing an intuitive interface for data review and export.

NEW FEATURE: Auto-populate Testing Name field using fuzzy matching against lookup table
"""

import os
import re
import io
import json
import time
import base64
from typing import List, Dict, Any, Tuple

import streamlit as st
import pandas as pd
import requests

# NEW: Fuzzy matching library
try:
    from rapidfuzz import fuzz
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False

# ---------------------------
# PDF Extraction Imports
# ---------------------------
HAVE_PYMUPDF = False
HAVE_PDFPLUMBER = False

try:
    import fitz  # PyMuPDF
    HAVE_PYMUPDF = True
except Exception:
    try:
        import pdfplumber
        HAVE_PDFPLUMBER = True
    except Exception:
        pass

# ---------------------------
# Excel Export
# ---------------------------
try:
    from openpyxl.styles import Font, PatternFill, Alignment
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

# ---------------------------
# API Configuration (Backend)
# ---------------------------
API_CONFIG = {
    'client_id': '074c933c-112f-4acf-a6a5-3199e4c78eea',
    'client_secret': 'ff7c6a75-1336-4594-b74e-f26065b87d4e',
    'model_name': 'gpt-4.1',
    'token_url': 'https://api-gw.boehringer-ingelheim.com:443/api/oauth/token',
    'api_url': 'https://api-gw.boehringer-ingelheim.com:443/apollo/llm-api/',
    'temperature': 0.2,
    'max_tokens': 10000,
    'completions_path': 'chat/completions'
}

# ---------------------------
# Column Schema (13 Columns)
# ---------------------------
ALL_COLUMNS = [
    "Testing Name",
    "Year",
    "Lab Code",
    "Farm/House/Flock code",
    "Country",
    "Types of bird",
    "Sample Size",
    "Disease",
    "Age (week)",
    "Test kit",
    "GMT",
    "Mean",
    "%CV",
]

# AI-Extracted columns (for validation)
AI_COLUMNS = [
    "Year", "Lab Code", "Farm/House/Flock code", "Country",
    "Types of bird", "Sample Size", "Disease", "Age (week)",
    "Test kit", "GMT", "Mean", "%CV"
]

# Country mapping for 2-letter codes
COUNTRY_MAPPING = {
    "philippines": "PH", "phil": "PH", "ph": "PH",
    "thailand": "TH", "thai": "TH", "th": "TH",
    "indonesia": "ID", "indo": "ID", "id": "ID",
    "malaysia": "MY", "malay": "MY", "my": "MY",
    "singapore": "SG", "sing": "SG", "sg": "SG",
    "vietnam": "VN", "viet nam": "VN", "vn": "VN",
    "myanmar": "MM", "burma": "MM", "mm": "MM",
    "cambodia": "KH", "kh": "KH",
    "laos": "LA", "lao": "LA", "la": "LA",
    "brunei": "BN", "bn": "BN",
    "timor-leste": "TL", "timor": "TL", "tl": "TL",
}

# ---------------------------
# NEW: Load Testing Name Lookup Table
# ---------------------------
LOOKUP_DF = None
LOOKUP_FILE = "testing_name_lookup.xlsx"

def load_lookup_table():
    """Load the Testing Name lookup table from Excel file."""
    global LOOKUP_DF
    try:
        if os.path.exists(LOOKUP_FILE):
            LOOKUP_DF = pd.read_excel(LOOKUP_FILE)
            # Clean up any NaN values to empty strings
            LOOKUP_DF = LOOKUP_DF.fillna("")
            return True
        else:
            st.warning(f"⚠️ Lookup file '{LOOKUP_FILE}' not found. Testing Name auto-population disabled.")
            return False
    except Exception as e:
        st.warning(f"⚠️ Could not load lookup file: {str(e)}. Testing Name auto-population disabled.")
        return False

# ---------------------------
# LLM Prompts
# ---------------------------
SYSTEM_PROMPT = """You are a precise data-extraction assistant for veterinary lab reports.
Extract all available records that match the target schema from the provided document text.

IMPORTANT FIELD-SPECIFIC RULES:
1. Year: Extract from test/bleeding date (e.g., "07/23/2025" → 2025)
2. Lab Code: Extract from "Lab code" or "Test Request No."
3. Farm/House/Flock code: Extract from "Customer-Name" or "Name of Client"
4. Country: ALWAYS use 2-letter ISO codes (PH, TH, ID, MY, SG, VN, MM, KH, LA, BN, TL)
   - If you see "Philippines" or location in Philippines → "PH"
   - If you see "Thailand" → "TH", etc.
5. Types of bird: Extract from "Type" field (e.g., "Commercial Layer", "Broiler")
6. Sample Size: Extract "No. Samples" value
7. Disease: Extract from "Assay" field and use SHORT CODES:
   - NDV → "ND" or "NDV"
   - IBV → "IB" or "IBV"
   - IBD → "IBD" or "IBDV"
   - ART → "ART" or "aMPV"
8. Age (week): Extract from "Age" field (e.g., "77 Week(s)" → 77)
9. Test kit: Extract manufacturer from "Reference" section or assay details (e.g., "Biochek", "ID Vet", "HIPRA")
10. GMT: Extract "GMT" value (geometric mean titer)
11. Mean: Extract "Mean Titer" value
12. %CV: Extract "%CV" value (without % symbol, as number)

CRITICAL: This PDF may contain MULTIPLE disease tests (ND, IB, IBD, ART).
Create ONE RECORD PER DISEASE TEST with shared metadata (farm, age, etc.) but different disease-specific values.

Return ONLY a valid JSON object:
{
  "records": [
    {
      "Year": int or null,
      "Lab Code": string or null,
      "Farm/House/Flock code": string or null,
      "Country": string (2-letter code) or null,
      "Types of bird": string or null,
      "Sample Size": int or null,
      "Disease": string or null,
      "Age (week)": int or null,
      "Test kit": string or null,
      "GMT": float or int or null,
      "Mean": float or int or null,
      "%CV": float or int or null
    }
  ]
}

Rules:
- Use numbers for numeric fields (no % sign in %CV)
- If a field is missing/unknown, use null
- Extract ALL disease tests as separate records
- Do NOT invent data not found in the text
- Output strictly as JSON. No comments, no explanation.
"""

USER_EXTRA_INSTRUCTION = """Extract data from this veterinary lab report into the following format:

Columns: Year | Lab Code | Farm/House/Flock code | Country | Types of bird | Sample Size | Disease | Age (week) | Test kit | GMT | Mean | %CV

IMPORTANT:
- Country must be 2-letter code (PH, TH, ID, MY, SG, VN, MM, KH, LA, BN, TL)
- Disease codes: NDV→ND or NDV, IBV→IB or IBV, IBD→IBD or IBDV, ART→ART or aMPV
- Create ONE row per disease test found in the document
- Extract actual numeric values for GMT, Mean, %CV (no % symbols)

Example row format (for reference only):
2025 | 15753 | LABASTIDA FARM (BUILDING 3) | PH | Commercial Layer | 10 | ND | 77 | Biochek | 5910 | 6164 | 31
"""

# ---------------------------
# Streamlit Page Config
# ---------------------------
st.set_page_config(
    page_title="AH Poultry Antibody Titer Dashboard - AI Data Extractor",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------
# Custom CSS Styling
# ---------------------------
st.markdown("""
<style>
    .main {
        background-color: white;
    }
    .stApp {
        background-color: white !important;
    }

    /* Title styling */
    .dashboard-title {
        text-align: center;
        color: #08312A;
        font-size: 22px;
        font-weight: bold;
        font-family: 'Verdana', sans-serif;
        margin-bottom: 5px;
    }
    .dashboard-subtitle {
        text-align: center;
        color: #6b7280;
        font-size: 13px;
        font-family: 'Verdana', sans-serif;
        margin-bottom: 20px;
    }

    /* Section headers */
    .section-header {
        font-size: 14px;
        font-weight: bold;
        color: #08312A;
        margin: 15px 0 10px 0;
        padding-bottom: 5px;
        border-bottom: 2px solid #00E47C;
    }

    /* Legend bar */
    .legend-bar {
        background: linear-gradient(90deg, #f0f9ff 0%, #d1fae5 25%, #fef3c7 50%, #fed7aa 75%);
        padding: 12px 20px;
        border-radius: 8px;
        margin: 15px 0;
        font-size: 13px;
        font-weight: 600;
        border: 1px solid #e5e7eb;
        text-align: center;
    }

    /* Status messages */
    .status-box {
        padding: 12px;
        border-radius: 8px;
        margin: 10px 0;
        font-size: 13px;
    }
    .status-success {
        background-color: #d1fae5;
        border: 1px solid #00E47C;
        color: #065f46;
    }
    .status-warning {
        background-color: #fef3c7;
        border: 1px solid #f59e0b;
        color: #92400e;
    }
    .status-info {
        background-color: #dbeafe;
        border: 1px solid #3b82f6;
        color: #1e40af;
    }

    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, #00E47C 0%, #08312A 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 10px 20px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 15px rgba(0, 228, 124, 0.3) !important;
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(0, 228, 124, 0.4) !important;
    }

    /* Sidebar styling */
    section[data-testid="stSidebar"] {
        background-color: #08312A;
        width: 280px !important;
    }
    section[data-testid="stSidebar"] * {
        color: white !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255, 255, 255, 0.2) !important;
    }
    [data-testid="collapsedControl"] {
        display: none;
    }

    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    /* Data editor styling */
    .stDataFrame {
        font-size: 12px;
    }
    
    /* Color empty cells with light yellow background */
    div[data-testid="stDataFrame"] div[data-testid="stDataFrameCell"] input[value=""],
    div[data-testid="stDataFrame"] div[data-testid="stDataFrameCell"] textarea:empty {
        background-color: #fef3c7 !important;
    }
    
    /* When user clicks to edit, make it white */
    div[data-testid="stDataFrame"] div[data-testid="stDataFrameCell"] input[value=""]:focus,
    div[data-testid="stDataFrame"] div[data-testid="stDataFrameCell"] textarea:empty:focus {
        background-color: #ffffff !important;
    }

    /* PDF preview */
    .pdf-preview-container {
        border: 2px solid #e5e7eb;
        border-radius: 8px;
        padding: 10px;
        background-color: #f9fafb;
    }
    
    /* NEW: Match confidence styling */
    .match-confidence {
        font-size: 12px;
        padding: 8px 12px;
        margin: 4px 0;
        border-radius: 6px;
        border-left: 4px solid;
    }
    .match-high {
        background-color: #d1fae5;
        border-color: #00E47C;
        color: #065f46;
    }
    .match-medium {
        background-color: #fef3c7;
        border-color: #f59e0b;
        color: #92400e;
    }
    .match-low {
        background-color: #fee2e2;
        border-color: #ef4444;
        color: #991b1b;
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------
# Utility Functions
# ---------------------------


def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    """Extract text from PDF bytes using available library."""
    texts = []

    if HAVE_PYMUPDF:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text") or ""
            if text.strip():
                texts.append(f"\n--- PAGE {page_num+1} ---\n{text}")
        doc.close()
        return "\n".join(texts)

    if HAVE_PDFPLUMBER:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = (page.extract_text() or "").strip()
                if text:
                    texts.append(f"\n--- PAGE {i} ---\n{text}")
        return "\n".join(texts)

    raise RuntimeError("No PDF library available. Install PyMuPDF: pip install PyMuPDF")


def safe_json_loads(content: str) -> Dict[str, Any]:
    """Parse JSON that may be wrapped in code fences."""
    txt = content.strip()
    txt = re.sub(r"^```(?:json)?\s*", "", txt)
    txt = re.sub(r"\s*```$", "", txt)
    return json.loads(txt)


def normalize_country(country_str):
    """Convert country name to 2-letter code."""
    if not country_str or pd.isna(country_str):
        return ""

    # Already 2-letter code
    if len(str(country_str)) == 2 and str(country_str).isalpha():
        return str(country_str).upper()

    # Map from full name
    country_lower = str(country_str).lower().strip()
    return COUNTRY_MAPPING.get(country_lower, "")


# ---------------------------
# NEW: Testing Name Matching Function
# ---------------------------

def find_testing_name(extracted_data: Dict[str, Any], pdf_text: str) -> Tuple[str, int, str, List[Dict]]:
    """
    Find the best matching Testing Name using multi-level fuzzy matching.
    
    Args:
        extracted_data: Dictionary with extracted fields (Disease, Test kit, etc.)
        pdf_text: Full text extracted from PDF for keyword searching
        
    Returns:
        Tuple of (testing_name, confidence_score, match_details, alternatives)
    """
    if LOOKUP_DF is None or not HAVE_RAPIDFUZZ:
        return "", 0, "Lookup table not available", []
    
    disease = str(extracted_data.get("Disease", "")).strip().upper()
    test_kit = str(extracted_data.get("Test kit", "")).strip()
    
    # Normalize disease codes
    disease_map = {
        "ND": "NDV", "NDV": "NDV",
        "IB": "IBV", "IBV": "IBV", 
        "IBD": "IBDV", "IBDV": "IBDV",
        "ART": "aMPV", "AMPV": "aMPV", "METAPNEUMOVIRUS": "aMPV",
        "ILT": "ILT",
        "H9": "H9"
    }
    disease = disease_map.get(disease, disease)
    
    if not disease:
        return "", 0, "No disease field found", []
    
    # Filter lookup table by disease (mandatory match)
    disease_matches = LOOKUP_DF[LOOKUP_DF['Disease'].str.upper() == disease]
    
    if len(disease_matches) == 0:
        return "", 0, f"No matches found for disease: {disease}", []
    
    # Score each potential match
    scored_matches = []
    
    for idx, row in disease_matches.iterrows():
        score = 0
        match_details = []
        
        # 1. Brand/Test Kit Match (50 points)
        brand = str(row.get('Brand', '')).strip()
        if brand and test_kit:
            brand_similarity = fuzz.ratio(brand.lower(), test_kit.lower())
            if brand_similarity >= 90:
                score += 50
                match_details.append(f"Brand exact match: {brand}")
            elif brand_similarity >= 80:
                score += 35
                match_details.append(f"Brand fuzzy match: {brand} ({brand_similarity}%)")
            elif brand.lower() in test_kit.lower() or test_kit.lower() in brand.lower():
                score += 20
                match_details.append(f"Brand partial match: {brand}")
        
        # 2. Type of Test Match (20 points)
        test_type = str(row.get('Type_of_test', '')).strip().upper()
        if test_type:
            if test_type in pdf_text.upper():
                type_similarity = fuzz.partial_ratio(test_type, pdf_text.upper())
                if type_similarity >= 90:
                    score += 20
                    match_details.append(f"Test type found: {test_type}")
                elif type_similarity >= 80:
                    score += 12
                    match_details.append(f"Test type fuzzy found: {test_type}")
        
        # 3. "How to Identify" Keywords (30 points total)
        keyword_score = 0
        keywords_found = []
        for i in range(1, 4):
            keyword = str(row.get(f'how_to_identify_{i}', '')).strip()
            if keyword and keyword in pdf_text:
                if i == 1:
                    keyword_score += 12
                elif i == 2:
                    keyword_score += 10
                else:
                    keyword_score += 8
                keywords_found.append(keyword)
        
        score += keyword_score
        if keywords_found:
            match_details.append(f"Keywords found: {', '.join(keywords_found)}")
        
        # Store this match
        scored_matches.append({
            'testing_name': row['Testing_Name'],
            'score': score,
            'details': match_details,
            'brand': brand,
            'test_type': test_type
        })
    
    # Sort by score
    scored_matches.sort(key=lambda x: x['score'], reverse=True)
    
    if len(scored_matches) == 0:
        return "", 0, "No matches found", []
    
    # Get best match
    best_match = scored_matches[0]
    testing_name = best_match['testing_name']
    confidence = best_match['score']
    details = " | ".join(best_match['details']) if best_match['details'] else "Disease match only"
    
    # Get alternatives (if any)
    alternatives = scored_matches[1:4] if len(scored_matches) > 1 else []
    
    return testing_name, confidence, details, alternatives


def normalize_record(rec: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize Extracted record to match schema."""
    # Start with AI-Extracted columns (B-M)
    out = {
        "Testing Name": "",
        "Year": rec.get("Year"),
        "Lab Code": rec.get("Lab Code"),
        "Farm/House/Flock code": rec.get("Farm/House/Flock code"),
        "Country": normalize_country(rec.get("Country")),
        "Types of bird": rec.get("Types of bird"),
        "Sample Size": rec.get("Sample Size"),
        "Disease": rec.get("Disease"),
        "Age (week)": rec.get("Age (week)"),
        "Test kit": rec.get("Test kit"),
        "GMT": rec.get("GMT"),
        "Mean": rec.get("Mean"),
        "%CV": rec.get("%CV"),
    }

    # Convert all AI fields to strings to avoid type conflicts
    # Year - convert to string
    if out["Year"] not in (None, ""):
        try:
            out["Year"] = str(int(str(out["Year"]).strip()))
        except Exception:
            out["Year"] = ""
    else:
        out["Year"] = ""

    # Sample Size - convert to string
    if out["Sample Size"] not in (None, ""):
        try:
            out["Sample Size"] = str(int(str(out["Sample Size"]).strip()))
        except Exception:
            out["Sample Size"] = ""
    else:
        out["Sample Size"] = ""

    # Age (week) - convert to string
    if out["Age (week)"] not in (None, ""):
        try:
            out["Age (week)"] = str(int(str(out["Age (week)"]).strip()))
        except Exception:
            out["Age (week)"] = ""
    else:
        out["Age (week)"] = ""

    # Numeric fields with decimals - convert to string
    for nf in ["%CV", "GMT", "Mean"]:
        if out[nf] not in (None, ""):
            try:
                sval = str(out[nf]).strip().replace("%", "")
                out[nf] = str(float(sval))
            except Exception:
                out[nf] = ""
        else:
            out[nf] = ""

    # String fields - replace None with empty string
    str_fields = ["Lab Code", "Farm/House/Flock code", "Types of bird", "Disease", "Test kit"]
    for sf in str_fields:
        if out[sf] is not None and out[sf] != "":
            out[sf] = str(out[sf]).strip()
        else:
            out[sf] = ""

    return out


def get_access_token(client_id: str, client_secret: str, token_url: str) -> str:
    """Get OAuth access token."""
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret,
    }
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    resp = requests.post(token_url, data=data, headers=headers, timeout=30)

    if not resp.ok:
        raise RuntimeError(f"Token request failed ({resp.status_code}): {resp.text}")

    js = resp.json()
    token = js.get("access_token")
    if not token:
        raise RuntimeError(f"No access_token in response: {js}")
    return token


def call_llm_api(access_token: str, user_prompt: str) -> Dict[str, Any]:
    """Call LLM API for data extraction."""
    url = API_CONFIG['api_url'].rstrip('/') + '/' + API_CONFIG['completions_path'].lstrip('/')

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": API_CONFIG['model_name'],
        "temperature": API_CONFIG['temperature'],
        "max_tokens": API_CONFIG['max_tokens'],
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        "response_format": {"type": "json_object"}
    }

    resp = requests.post(url, headers=headers, json=payload, timeout=120)

    if not resp.ok:
        raise RuntimeError(f"LLM request failed ({resp.status_code}): {resp.text}")

    js = resp.json()
    content = js["choices"][0]["message"]["content"]
    return safe_json_loads(content)


def extract_from_pdf(pdf_bytes: bytes, filename: str, access_token: str) -> Tuple[List[Dict], str, bool, List[Dict]]:
    """
    Extract data from single PDF file.
    Returns: (records, log_message, success_flag, match_info_list)
    """
    match_info_list = []  # NEW: Store matching information
    
    try:
        # Extract text
        text = extract_text_from_pdf_bytes(pdf_bytes)
        if not text.strip():
            # Create blank record for manual entry
            blank_record = {col: "" for col in ALL_COLUMNS}
            blank_record['_source_file'] = filename
            match_info_list.append({
                'filename': filename,
                'testing_name': '',
                'confidence': 0,
                'details': 'No text extracted',
                'alternatives': []
            })
            return [blank_record], f"⚠️ {filename}: No text extracted - Manual entry required", False, match_info_list

        # Call LLM
        prompt = f"Document: {filename}\n\n{USER_EXTRA_INSTRUCTION}\n\n{text}"
        data = call_llm_api(access_token, prompt)

        raw_records = data.get("records", [])
        if not isinstance(raw_records, list) or len(raw_records) == 0:
            # Create blank record for manual entry
            blank_record = {col: "" for col in ALL_COLUMNS}
            blank_record['_source_file'] = filename
            match_info_list.append({
                'filename': filename,
                'testing_name': '',
                'confidence': 0,
                'details': 'No data extracted',
                'alternatives': []
            })
            return [blank_record], f"⚠️ {filename}: No data extracted - Manual entry required", False, match_info_list

        # Normalize records
        normalized = [normalize_record(rec) for rec in raw_records]
        
        # NEW: Auto-populate Testing Name using fuzzy matching
        for rec in normalized:
            testing_name, confidence, details, alternatives = find_testing_name(rec, text)
            
            # Only auto-fill if confidence >= 70%
            if confidence >= 70:
                rec['Testing Name'] = testing_name
            
            # Store match info for display
            match_info_list.append({
                'filename': filename,
                'disease': rec.get('Disease', ''),
                'test_kit': rec.get('Test kit', ''),
                'testing_name': testing_name,
                'confidence': confidence,
                'details': details,
                'alternatives': alternatives
            })

        return normalized, f"✅ {filename}: {len(normalized)} record(s) extracted", True, match_info_list

    except Exception as e:
        # Create blank record for manual entry even on error
        blank_record = {col: "" for col in ALL_COLUMNS}
        blank_record['_source_file'] = filename
        match_info_list.append({
            'filename': filename,
            'testing_name': '',
            'confidence': 0,
            'details': f'Error: {str(e)}',
            'alternatives': []
        })
        return [blank_record], f"❌ {filename}: Extraction failed - Manual entry required ({str(e)})", False, match_info_list


def create_excel_download(df: pd.DataFrame, include_audit: bool = False) -> bytes:
    """
    Create Excel file with formatting (clean column names, including Source File).
    
    Args:
        df: DataFrame to export
        include_audit: If True, add a second sheet with match confidence data
    """
    output = io.BytesIO()

    # Create clean dataframe without emojis
    download_df = df.copy()

    # Clean column names (remove emojis)
    clean_columns = {}
    for col in download_df.columns:
        # Remove emoji prefixes
        clean_col = col
        clean_col = clean_col.replace("📄 ", "").replace("🔍 ", "").replace("✏️ ", "")
        clean_columns[col] = clean_col

    download_df = download_df.rename(columns=clean_columns)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        download_df.to_excel(writer, index=False, sheet_name='Lab Data')

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Lab Data']

        # Header formatting
        header_fill = PatternFill(start_color="00E47C", end_color="00E47C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # NEW: Add audit sheet if requested
        if include_audit and 'match_info' in st.session_state:
            audit_data = []
            for info in st.session_state.match_info:
                audit_data.append({
                    'Source File': info.get('filename', ''),
                    'Disease': info.get('disease', ''),
                    'Test Kit': info.get('test_kit', ''),
                    'Testing Name (Matched)': info.get('testing_name', ''),
                    'Confidence Score': info.get('confidence', 0),
                    'Match Details': info.get('details', ''),
                    'Alternatives': ', '.join([f"{alt['testing_name']} ({alt['score']}%)" 
                                              for alt in info.get('alternatives', [])])
                })
            
            if audit_data:
                audit_df = pd.DataFrame(audit_data)
                audit_df.to_excel(writer, index=False, sheet_name='Match Confidence')
                
                # Format audit sheet
                audit_sheet = writer.sheets['Match Confidence']
                for cell in audit_sheet[1]:
                    cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust audit sheet columns
                for column in audit_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    audit_sheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output.getvalue()


def display_pdf_preview(pdf_bytes: bytes):
    """Display PDF preview in iframe."""
    base64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
    pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="800" type="application/pdf"></iframe>'
    st.markdown(pdf_display, unsafe_allow_html=True)


def get_blank_ai_columns(df: pd.DataFrame) -> List[str]:
    """Get list of AI columns that have blank values."""
    blank_cols = []
    for col in AI_COLUMNS:
        if col in df.columns:
            # Check if column has any blank/empty values
            if df[col].isna().any() or (df[col] == "").any():
                blank_cols.append(col)
    return blank_cols


# ---------------------------
# Main Application
# ---------------------------

def main():
    # Load lookup table at startup
    load_lookup_table()
    
    # Sidebar - Logo and AH Poultry Image
    with st.sidebar:
        st.markdown("<br>", unsafe_allow_html=True)
        
        # BI Logo from Imgur
        bi_logo_url = "https://i.imgur.com/nAF4iae.png"
        st.markdown(
            f'<div style="text-align: center;"><img src="{bi_logo_url}" width="180" style="pointer-events: none;"></div>',
            unsafe_allow_html=True
        )
        
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        # AH Poultry from Imgur
        ah_poultry_url = "https://i.imgur.com/LumoN5n.png"
        st.markdown(
            f'<div style="text-align: center;"><img src="{ah_poultry_url}" width="200" style="pointer-events: none;"></div>',
            unsafe_allow_html=True
        )
        
        # NEW: Show lookup table status
        st.markdown("<br>", unsafe_allow_html=True)
        if LOOKUP_DF is not None:
            st.success(f"✅ Lookup table loaded\n({len(LOOKUP_DF)} testing names)")
        else:
            st.warning("⚠️ Lookup table not loaded")

    # Main Title
    st.markdown('<div class="dashboard-title">AH Poultry Antibody Titer Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="dashboard-subtitle">(Lab Reports - AI Data Extraction Utility)</div>', unsafe_allow_html=True)

    st.markdown("---")

    # Initialize session state
    if 'extraction_complete' not in st.session_state:
        st.session_state.extraction_complete = False
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = None
    if 'pdf_files' not in st.session_state:
        st.session_state.pdf_files = {}
    if 'show_download' not in st.session_state:
        st.session_state.show_download = False
    if 'match_info' not in st.session_state:  # NEW
        st.session_state.match_info = []

    # Step 1: Upload PDFs
    st.markdown('<div class="section-header">📄 Step 1: Upload Lab Reports</div>', unsafe_allow_html=True)

    uploaded_files = st.file_uploader(
        "Select one or more PDF lab reports",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload poultry lab reports in PDF format"
    )

    col1, col2, col3 = st.columns([1, 1, 2])

    with col1:
        extract_button = st.button("🚀 Extract Data", use_container_width=True, disabled=not uploaded_files)

    with col2:
        if st.session_state.extraction_complete:
            if st.button("🔄 Reset", use_container_width=True):
                st.session_state.extraction_complete = False
                st.session_state.extracted_data = None
                st.session_state.pdf_files = {}
                st.session_state.show_download = False
                st.session_state.match_info = []  # NEW
                st.rerun()

    # Process extraction
    if extract_button and uploaded_files:
        all_records = []
        pdf_storage = {}
        extraction_logs = []
        all_match_info = []  # NEW
        success_count = 0
        failed_count = 0

        with st.spinner("🔍 Processing lab reports..."):
            # Get access token
            try:
                access_token = get_access_token(
                    API_CONFIG['client_id'],
                    API_CONFIG['client_secret'],
                    API_CONFIG['token_url']
                )
            except Exception as e:
                st.error(f"❌ Authentication failed: {str(e)}")
                st.stop()

            # Process each PDF
            progress_bar = st.progress(0)
            for idx, uploaded_file in enumerate(uploaded_files):
                pdf_bytes = uploaded_file.read()
                pdf_storage[uploaded_file.name] = pdf_bytes

                records, log_msg, success, match_info = extract_from_pdf(pdf_bytes, uploaded_file.name, access_token)

                # Track success/failure
                if success:
                    success_count += 1
                else:
                    failed_count += 1
                
                extraction_logs.append(log_msg)
                all_match_info.extend(match_info)  # NEW

                # Add source file to each record
                for rec in records:
                    rec['_source_file'] = uploaded_file.name

                all_records.extend(records)

                progress_bar.progress((idx + 1) / len(uploaded_files))

            progress_bar.empty()

        # Always proceed if we have records (even blank ones)
        if all_records:
            # Display extraction summary
            if failed_count > 0:
                st.warning(f"⚠️ {failed_count} file(s) require manual entry. Please review and fill the data manually.")
            
            # Create DataFrame with all columns
            df = pd.DataFrame(all_records)

            # Ensure all columns exist in correct order
            for col in ALL_COLUMNS:
                if col not in df.columns:
                    df[col] = ""

            # Add source file column at the end
            df['Source File'] = df['_source_file']
            df = df.drop('_source_file', axis=1)

            # Reorder columns: Source File first, then all others
            df = df[['Source File'] + ALL_COLUMNS]

            st.session_state.extracted_data = df
            st.session_state.pdf_files = pdf_storage
            st.session_state.match_info = all_match_info  # NEW
            st.session_state.extraction_complete = True
            
            success_msg = f"✅ Extraction complete: {len(df)} record(s) from {len(uploaded_files)} file(s)"
            if failed_count > 0:
                success_msg += f" ({success_count} successful, {failed_count} require manual entry)"
            st.success(success_msg)
        else:
            st.error("❌ Unexpected error: No records created. Please try again.")

    # Step 2: Review and Edit Data
    if st.session_state.extraction_complete and st.session_state.extracted_data is not None:
        st.markdown("---")
        st.markdown('<div class="section-header">✏️ Step 2: Review & Edit Extracted Data</div>', unsafe_allow_html=True)

        df = st.session_state.extracted_data.copy()

        # Summary info
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f'<div class="status-box status-info">📊 <strong>{len(df)}</strong> records extracted</div>', unsafe_allow_html=True)
        with col2:
            missing_count = (df['Testing Name'] == '').sum()
            st.markdown(f'<div class="status-box status-warning">⚠️ <strong>{missing_count}</strong> records need Testing Name</div>', unsafe_allow_html=True)
        with col3:
            unique_files = df['Source File'].nunique()
            st.markdown(f'<div class="status-box status-success">📄 <strong>{unique_files}</strong> source file(s)</div>', unsafe_allow_html=True)

        # NEW: Display AI Match Confidence Summary
        if st.session_state.match_info and LOOKUP_DF is not None:
            with st.expander("🔍 AI Match Confidence Summary (Testing Name Auto-Population)", expanded=False):
                for info in st.session_state.match_info:
                    confidence = info.get('confidence', 0)
                    testing_name = info.get('testing_name', '')
                    details = info.get('details', '')
                    filename = info.get('filename', '')
                    disease = info.get('disease', '')
                    test_kit = info.get('test_kit', '')
                    alternatives = info.get('alternatives', [])
                    
                    # Determine confidence level
                    if confidence >= 90:
                        conf_class = "match-high"
                        conf_icon = "✅"
                        conf_label = "High Confidence"
                    elif confidence >= 70:
                        conf_class = "match-medium"
                        conf_icon = "⚠️"
                        conf_label = "Medium Confidence"
                    else:
                        conf_class = "match-low"
                        conf_icon = "❓"
                        conf_label = "Low/No Match"
                    
                    # Display match info
                    st.markdown(f"""
                    <div class="match-confidence {conf_class}">
                        <strong>{conf_icon} {filename}</strong><br>
                        Disease: {disease} | Test Kit: {test_kit}<br>
                        <strong>{conf_label} ({confidence}%):</strong> {testing_name if testing_name else "No match - manual entry required"}<br>
                        <small>{details}</small>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Show alternatives if any
                    if alternatives and len(alternatives) > 0:
                        alt_text = " | ".join([f"{alt['testing_name']} ({alt['score']}%)" for alt in alternatives[:2]])
                        st.markdown(f"<small style='color: #6b7280; margin-left: 20px;'>Alternatives: {alt_text}</small>", unsafe_allow_html=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)

        # Check for blank AI columns and display warning
        blank_ai_cols = get_blank_ai_columns(df)
        if blank_ai_cols:
            blank_cols_str = ", ".join(blank_ai_cols)
            st.markdown(f'<div class="status-box status-warning">⚠️ <strong>Missing data in field(s): </strong> {blank_cols_str}<br><small>Either information is missing in lab report (or) could not be extracted. <br>Please review concerned lab report manually to fill respective column (or) leave them as blank.</small></div>', unsafe_allow_html=True)
        
        # Check if any records are completely blank (failed extraction)
        blank_records = df[df[AI_COLUMNS].apply(lambda row: all(row == ''), axis=1)]
        if len(blank_records) > 0:
            blank_files = blank_records['Source File'].tolist()
            files_list = "<br>".join([f"• {f}" for f in blank_files])
            st.markdown(f'<div class="status-box status-warning">📝 <strong>Manual Record Entry Required:</strong> Below lab report(s) could not be auto-extracted.<br><small>{files_list}</small></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Compact legend bar
        st.markdown("""
        <div style="text-align: center; font-size: 12px; color: #6b7280; margin: 8px 0 5px 0; padding: 5px;">
            📄 Source File(s) &nbsp;|&nbsp; 🔍 AI-Extracted (editable) &nbsp;|&nbsp; ✏️ Testing Name (Auto-populated or Manual)
            <br><small style="color: #9ca3af; font-size: 11px;">Please Click "💾 Save Changes" to download Excel with latest information</small>
        </div>
        """, unsafe_allow_html=True)

        # Create display dataframe
        display_df = df.copy()
        
        # Clean up any None or NaN values to empty strings
        if 'Testing Name' in display_df.columns:
            display_df['Testing Name'] = display_df['Testing Name'].apply(
                lambda x: "" if pd.isna(x) else str(x).strip()
            )
        
        for col in AI_COLUMNS:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(
                    lambda x: "" if pd.isna(x) else str(x).strip()
                )

        # Add emoji prefixes to column names
        display_columns = {}
        for col in display_df.columns:
            if col == "Source File":
                display_columns[col] = "📄 Source File"
            elif col == "Testing Name":
                display_columns[col] = "✏️ Testing Name"
            elif col in AI_COLUMNS:
                display_columns[col] = f"🔍 {col}"
            else:
                display_columns[col] = f"✏️ {col}"

        display_df = display_df.rename(columns=display_columns)

        # Configure column types for data editor
        column_config = {
            "📄 Source File": st.column_config.TextColumn(
                "📄 Source File",
                help="Source PDF file (editable)"
            ),
            "✏️ Testing Name": st.column_config.TextColumn(
                "✏️ Testing Name",
                help="⚠️ MANDATORY - Auto-populated or enter manually",
                required=True
            ),
            "🔍 Year": st.column_config.TextColumn(
                "🔍 Year",
                help="AI-Extracted Year (editable if needed)"
            ),
            "🔍 Lab Code": st.column_config.TextColumn(
                "🔍 Lab Code",
                help="AI-Extracted Lab Code (editable if needed)"
            ),
            "🔍 Farm/House/Flock code": st.column_config.TextColumn(
                "🔍 Farm/House/Flock code",
                help="AI-Extracted Farm/House/Flock Code (editable if needed)"
            ),
            "🔍 Country": st.column_config.TextColumn(
                "🔍 Country",
                help="AI-Extracted Country - 2-letter code (editable if needed)"
            ),
            "🔍 Types of bird": st.column_config.TextColumn(
                "🔍 Types of bird",
                help="AI-Extracted Bird Type (editable if needed)"
            ),
            "🔍 Sample Size": st.column_config.TextColumn(
                "🔍 Sample Size",
                help="AI-Extracted Sample Size (editable if needed)"
            ),
            "🔍 Disease": st.column_config.TextColumn(
                "🔍 Disease",
                help="AI-Extracted Disease Code (editable if needed)"
            ),
            "🔍 Age (week)": st.column_config.TextColumn(
                "🔍 Age (week)",
                help="AI-Extracted Age in Weeks (editable if needed)"
            ),
            "🔍 Test kit": st.column_config.TextColumn(
                "🔍 Test kit",
                help="AI-Extracted Test Kit (editable if needed)"
            ),
            "🔍 GMT": st.column_config.TextColumn(
                "🔍 GMT",
                help="AI-Extracted GMT Value (editable if needed)"
            ),
            "🔍 Mean": st.column_config.TextColumn(
                "🔍 Mean",
                help="AI-Extracted Mean Value (editable if needed)"
            ),
            "🔍 %CV": st.column_config.TextColumn(
                "🔍 %CV",
                help="AI-Extracted %CV Value (editable if needed)"
            ),
        }

        # Editable data table
        edited_df = st.data_editor(
            display_df,
            column_config=column_config,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            height=250,
            key="main_data_editor"
        )

        # Check if all Testing Names are filled
        testing_names_filled = True
        if '✏️ Testing Name' in edited_df.columns:
            empty_count = (edited_df['✏️ Testing Name'] == '').sum()
            testing_names_filled = (empty_count == 0)
        
        # Save Changes Button
        st.markdown("<br>", unsafe_allow_html=True)
        
        if not testing_names_filled:
            st.warning(f"⚠️ 'Testing Name' is Missing")
        
        col_save1, col_save2 = st.columns([1, 3])
        
        with col_save1:
            if st.button("💾 Save Changes", use_container_width=True, type="primary", disabled=not testing_names_filled):
                # Create clean version (remove emoji prefixes only)
                download_df = edited_df.copy()
                clean_columns = {}
                for col in download_df.columns:
                    clean_col = col.replace("📄 ", "").replace("🔍 ", "").replace("✏️ ", "")
                    clean_columns[col] = clean_col
                download_df = download_df.rename(columns=clean_columns)
                
                # Save to session state
                st.session_state.extracted_data = download_df
                st.session_state.show_download = True
                st.success("✅ Changes saved! Scroll down to download.")
                st.rerun()

        # PDF Preview
        st.markdown("---")
        st.markdown('<div class="section-header">📑 PDF Previewer</div>', unsafe_allow_html=True)

        with st.expander("Click to view source PDFs", expanded=False):
            if st.session_state.pdf_files:
                pdf_names = list(st.session_state.pdf_files.keys())
                selected_pdf = st.selectbox("Select PDF to preview:", pdf_names)

                if selected_pdf:
                    st.markdown(f"**Viewing:** {selected_pdf}")
                    display_pdf_preview(st.session_state.pdf_files[selected_pdf])
            else:
                st.info("No PDFs available for preview")

        # Step 3 - Only show after saving
        if 'show_download' in st.session_state and st.session_state.show_download:
            # Step 3: Download
            st.markdown("---")
            st.markdown('<div class="section-header">📥 Step 3: Download Excel File</div>', unsafe_allow_html=True)

            # Use the saved data from session state
            download_ready_df = st.session_state.extracted_data
            
            # Check if Testing Name is filled
            if 'Testing Name' in download_ready_df.columns:
                missing_testing_names = (download_ready_df['Testing Name'] == '').sum()
            else:
                missing_testing_names = len(download_ready_df)

            if missing_testing_names > 0:
                st.markdown(f'<div class="status-box status-warning">⚠️ Please enter "Testing Name" for all {missing_testing_names} record(s) before downloading</div>', unsafe_allow_html=True)
                st.button("📥 Download Excel", disabled=True, use_container_width=True)
            else:
                st.markdown('<div class="status-box status-success">✅ Ready to download!</div>', unsafe_allow_html=True)

                # NEW: Download format selector
                col_format1, col_format2 = st.columns([1, 2])
                with col_format1:
                    download_format = st.radio(
                        "Download Format:",
                        ["Clean Data Only", "With Match Confidence (Audit Trail)"],
                        help="Choose whether to include match confidence data in a separate sheet"
                    )
                
                include_audit = (download_format == "With Match Confidence (Audit Trail)")

                # Prepare Excel file
                excel_bytes = create_excel_download(download_ready_df, include_audit=include_audit)

                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_bytes,
                    file_name=f"Poultry_Lab_Data_{time.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


if __name__ == "__main__":
    main()